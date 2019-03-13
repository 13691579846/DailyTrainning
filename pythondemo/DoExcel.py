from openpyxl import load_workbook
# 读取excel文件中的数据
class DoExcel(object):
    def __init__(self, filePath, sheetName):
        self.filePath = filePath
        self.sheetName = sheetName
    # 读数据
    def do_exlce(self):
        wd = load_workbook(self.filePath) # 打开excel文件
        sh = wd[self.sheetName] # 获取sheetname表单的对象

        testData=[]
        for i in range(2, sh.max_row+1):
            sub_data = {}
            sub_data['id'] = sh.cell(i, 1).value # 获取第2行第1列的数据·
            sub_data['tcTitle'] = sh.cell(i, 2).value
            sub_data['method'] = sh.cell(i, 3).value
            sub_data['url'] = sh.cell(i, 4).value
            sub_data['param'] = sh.cell(i, 5).value
            testData.append(sub_data)
        return testData
    # 写数据
    def writeData(self, row, colum, value):
        wd = load_workbook(self.filePath)
        sh = wd[self.sheetName]

        sh.cell(row, colum).value = value

        wd.save(self.filePath)
if __name__ == '__main__':
    doexcel = DoExcel('./dataexcel.xlsx','login')
    data =doexcel.do_exlce()
    print(doexcel)
    for value in data:
        print(value)
    doexcel.writeData(2, 6, 'value')